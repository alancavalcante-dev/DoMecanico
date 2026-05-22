from django.db.models import Q, F, Count, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes, throttle_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle


class OSPublicaBuscarThrottle(AnonRateThrottle):
    rate = '5/minute'
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import io

from .models import (
    Cliente, Veiculo, FotoVeiculo, Funcionario,
    Peca, MovimentacaoEstoque, OrdemServico,
    ServicoOS, PecaOS, NotaFiscal,
    ChecklistEntrada, DanoChecklist,
    Agendamento, Orcamento, ItemOrcamento,
    GarantiaServico, GarantiaDefault, ComissaoMecanico, AlertaEstoque,
)
from .serializers import (
    ClienteSerializer, ClienteListSerializer,
    VeiculoSerializer, VeiculoListSerializer, FotoVeiculoSerializer,
    FuncionarioSerializer, PecaSerializer, MovimentacaoEstoqueSerializer,
    OrdemServicoSerializer, OrdemServicoListSerializer,
    ServicoOSSerializer, PecaOSSerializer, NotaFiscalSerializer,
    ChecklistEntradaSerializer, DanoChecklistSerializer, ChecklistPublicoSerializer,
    OSPublicaSerializer,
    AgendamentoSerializer, OrcamentoSerializer, ItemOrcamentoSerializer,
    OrcamentoPublicoSerializer, GarantiaServicoSerializer,
    ComissaoMecanicoSerializer, AlertaEstoqueSerializer,
)


def get_oficina(request):
    return request.user.membro.oficina


def _validar_imagem(arquivo, formatos_permitidos=('JPEG', 'PNG', 'WEBP', 'GIF')):
    """Verifica os bytes reais do arquivo — rejeita arquivos que não sejam imagens válidas."""
    from PIL import Image
    try:
        arquivo.seek(0)
        img = Image.open(io.BytesIO(arquivo.read(2048)))
        if img.format not in formatos_permitidos:
            return False
        arquivo.seek(0)
        return True
    except Exception:
        return False


class ClienteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return ClienteListSerializer
        return ClienteSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Cliente.objects.filter(oficina=oficina)
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(nome__icontains=search) | Q(cpf_cnpj__icontains=search) |
                Q(telefone__icontains=search) | Q(celular__icontains=search) |
                Q(email__icontains=search)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))


class VeiculoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return VeiculoListSerializer
        return VeiculoSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Veiculo.objects.select_related('cliente').filter(oficina=oficina)
        cliente_id = self.request.query_params.get('cliente')
        search = self.request.query_params.get('search')
        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)
        if search:
            qs = qs.filter(
                Q(placa__icontains=search) | Q(marca__icontains=search) |
                Q(modelo__icontains=search) | Q(cliente__nome__icontains=search)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))

    @action(detail=True, methods=['post'], url_path='upload-foto')
    def upload_foto(self, request, pk=None):
        veiculo = self.get_object()
        fotos = request.FILES.getlist('fotos')
        if not fotos:
            return Response({'erro': 'Nenhuma foto enviada.'}, status=status.HTTP_400_BAD_REQUEST)
        MAX_SIZE = 10 * 1024 * 1024  # 10 MB
        criadas = []
        for foto in fotos:
            if foto.size > MAX_SIZE:
                return Response({'erro': 'Arquivo muito grande. Limite: 10 MB.'}, status=status.HTTP_400_BAD_REQUEST)
            if not _validar_imagem(foto):
                return Response({'erro': 'Arquivo inválido. Envie uma imagem JPEG, PNG, WebP ou GIF.'}, status=status.HTTP_400_BAD_REQUEST)
            obj = FotoVeiculo.objects.create(veiculo=veiculo, foto=foto, descricao=request.data.get('descricao', ''))
            criadas.append(FotoVeiculoSerializer(obj, context={'request': request}).data)
        return Response(criadas, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path='foto/(?P<foto_id>[^/.]+)')
    def deletar_foto(self, request, pk=None, foto_id=None):
        try:
            foto = FotoVeiculo.objects.get(id=foto_id, veiculo_id=pk)
            foto.foto.delete()
            foto.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except FotoVeiculo.DoesNotExist:
            return Response({'erro': 'Foto não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'], url_path='saude')
    def saude(self, request, pk=None):
        veiculo = self.get_object()
        return _gerar_pdf_saude(veiculo)


class FuncionarioViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FuncionarioSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Funcionario.objects.filter(oficina=oficina)
        search = self.request.query_params.get('search')
        ativo = self.request.query_params.get('ativo')
        if search:
            qs = qs.filter(Q(nome__icontains=search) | Q(cpf__icontains=search) | Q(cargo__icontains=search))
        if ativo is not None:
            qs = qs.filter(ativo=ativo.lower() == 'true')
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))


class PecaViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = PecaSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Peca.objects.filter(oficina=oficina)
        search = self.request.query_params.get('search')
        estoque_baixo = self.request.query_params.get('estoque_baixo')
        if search:
            qs = qs.filter(Q(nome__icontains=search) | Q(codigo__icontains=search) | Q(marca__icontains=search))
        if estoque_baixo == 'true':
            qs = qs.filter(quantidade__lte=F('quantidade_minima'))
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))

    @action(detail=True, methods=['post'], url_path='movimentar')
    def movimentar(self, request, pk=None):
        peca = self.get_object()
        from decimal import Decimal
        tipo = request.data.get('tipo')
        quantidade = Decimal(str(request.data.get('quantidade', 0)))
        preco_unitario = Decimal(str(request.data.get('preco_unitario', 0)))
        motivo = request.data.get('motivo', '')

        if tipo not in ['entrada', 'saida', 'ajuste']:
            return Response({'erro': 'Tipo inválido.'}, status=status.HTTP_400_BAD_REQUEST)
        if quantidade <= 0:
            return Response({'erro': 'Quantidade deve ser maior que zero.'}, status=status.HTTP_400_BAD_REQUEST)

        if tipo == 'entrada':
            peca.quantidade += quantidade
        elif tipo == 'saida':
            if peca.quantidade < quantidade:
                return Response({'erro': 'Estoque insuficiente.'}, status=status.HTTP_400_BAD_REQUEST)
            peca.quantidade -= quantidade
        else:
            peca.quantidade = quantidade

        peca.save()
        mov = MovimentacaoEstoque.objects.create(
            peca=peca, tipo=tipo, quantidade=quantidade,
            preco_unitario=preco_unitario, motivo=motivo
        )
        # Gera alerta se estoque ficou baixo
        if peca.estoque_baixo:
            AlertaEstoque.objects.create(peca=peca, quantidade_no_alerta=peca.quantidade)
        return Response(MovimentacaoEstoqueSerializer(mov).data, status=status.HTTP_201_CREATED)


class MovimentacaoEstoqueViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = MovimentacaoEstoqueSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = MovimentacaoEstoque.objects.filter(peca__oficina=oficina).select_related('peca')
        peca_id = self.request.query_params.get('peca')
        if peca_id:
            qs = qs.filter(peca_id=peca_id)
        return qs


class OrdemServicoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return OrdemServicoListSerializer
        return OrdemServicoSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = OrdemServico.objects.filter(oficina=oficina).select_related('cliente', 'veiculo', 'mecanico')
        search = self.request.query_params.get('search')
        status_filter = self.request.query_params.get('status')
        cliente_id = self.request.query_params.get('cliente')
        if search:
            qs = qs.filter(
                Q(numero__icontains=search) | Q(cliente__nome__icontains=search) |
                Q(veiculo__placa__icontains=search)
            )
        if status_filter:
            qs = qs.filter(status=status_filter)
        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)
        mecanico_id = self.request.query_params.get('mecanico')
        if mecanico_id:
            qs = qs.filter(mecanico_id=mecanico_id)
        data_inicio = self.request.query_params.get('data_inicio')
        data_fim = self.request.query_params.get('data_fim')
        if data_inicio:
            qs = qs.filter(data_entrada__date__gte=data_inicio)
        if data_fim:
            qs = qs.filter(data_entrada__date__lte=data_fim)
        return qs

    def perform_create(self, serializer):
        oficina = get_oficina(self.request)
        ultimo = OrdemServico.objects.filter(oficina=oficina).order_by('-id').first()
        numero = f'OS{str((ultimo.id if ultimo else 0) + 1).zfill(6)}'
        serializer.save(oficina=oficina, numero=numero)

    @action(detail=True, methods=['post'], url_path='adicionar-servico')
    def adicionar_servico(self, request, pk=None):
        ordem = self.get_object()
        data = {
            'ordem': ordem.id,
            'descricao': request.data.get('descricao', ''),
            'quantidade': request.data.get('quantidade', 1),
            'preco_unitario': request.data.get('preco_unitario', 0),
        }
        serializer = ServicoOSSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='adicionar-peca')
    def adicionar_peca(self, request, pk=None):
        ordem = self.get_object()
        peca_id = request.data.get('peca') or None
        from decimal import Decimal
        quantidade = Decimal(str(request.data.get('quantidade', 1)))

        # Verifica estoque antes de qualquer coisa
        if peca_id:
            try:
                peca_obj = Peca.objects.get(id=peca_id, oficina=ordem.oficina)
                if peca_obj.quantidade < quantidade:
                    return Response({'erro': 'Estoque insuficiente.'}, status=status.HTTP_400_BAD_REQUEST)
            except Peca.DoesNotExist:
                return Response({'erro': 'Peça não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            peca_obj = None

        data = {
            'ordem': ordem.id,
            'peca': peca_id,
            'descricao': request.data.get('descricao', ''),
            'quantidade': quantidade,
            'preco_unitario': request.data.get('preco_unitario', 0),
        }
        serializer = PecaOSSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            if peca_obj:
                peca_obj.quantidade -= quantidade
                peca_obj.save()
                MovimentacaoEstoque.objects.create(
                    peca=peca_obj, tipo='saida', quantidade=quantidade,
                    preco_unitario=Decimal(str(request.data.get('preco_unitario', 0))),
                    motivo=f'OS {ordem.numero}'
                )
                if peca_obj.estoque_baixo:
                    AlertaEstoque.objects.create(peca=peca_obj, quantidade_no_alerta=peca_obj.quantidade)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['patch'], url_path='atualizar-status')
    def atualizar_status(self, request, pk=None):
        ordem = self.get_object()
        novo_status = request.data.get('status')
        if novo_status not in dict(OrdemServico.STATUS_CHOICES):
            return Response({'erro': 'Status inválido.'}, status=status.HTTP_400_BAD_REQUEST)
        ordem.status = novo_status
        if novo_status == 'concluida' and not ordem.data_conclusao:
            ordem.data_conclusao = timezone.now()
        ordem.save()
        # Gera comissão automaticamente se mecânico tem percentual configurado
        if novo_status == 'concluida' and ordem.mecanico and ordem.mecanico.percentual_comissao > 0:
            from decimal import Decimal
            base = Decimal(str(ordem.total_servicos))  # comissão só sobre mão de obra
            valor = base * ordem.mecanico.percentual_comissao / 100
            ComissaoMecanico.objects.update_or_create(
                funcionario=ordem.mecanico,
                ordem=ordem,
                defaults={'percentual': ordem.mecanico.percentual_comissao, 'valor': round(valor, 2)},
            )
        # Notificações ao concluir OS
        if novo_status == 'concluida':
            try:
                from .whatsapp import notificar_os_concluida
                notificar_os_concluida(ordem)
            except Exception:
                pass
            try:
                from .email_os import notificar_os_concluida_email
                notificar_os_concluida_email(ordem)
            except Exception:
                pass
        return Response(OrdemServicoSerializer(ordem).data)

    @action(detail=True, methods=['get'], url_path='gerar-pdf')
    def gerar_pdf(self, request, pk=None):
        ordem = self.get_object()
        return _gerar_pdf_os(ordem)


class ServicoOSViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ServicoOSSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        return ServicoOS.objects.filter(ordem__oficina=oficina)


class PecaOSViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = PecaOSSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        return PecaOS.objects.filter(ordem__oficina=oficina)


class NotaFiscalViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = NotaFiscalSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        return NotaFiscal.objects.filter(ordem__oficina=oficina).select_related('ordem', 'ordem__cliente')

    def create(self, request, *args, **kwargs):
        oficina = get_oficina(request)
        ordem_id = request.data.get('ordem')
        try:
            ordem = OrdemServico.objects.get(id=ordem_id, oficina=oficina)
        except OrdemServico.DoesNotExist:
            return Response({'erro': 'OS não encontrada.'}, status=400)
        if ordem.status != 'concluida':
            return Response({'erro': 'Só é possível emitir nota para OS com status Concluída.'}, status=400)
        if hasattr(ordem, 'nota_fiscal'):
            return Response({'erro': 'Esta OS já possui uma nota fiscal emitida.'}, status=400)
        ultima = NotaFiscal.objects.filter(ordem__oficina=oficina).order_by('-id').first()
        numero = f'NF{str((ultima.id if ultima else 0) + 1).zfill(6)}'
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(numero_nota=numero)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='imprimir')
    def imprimir(self, request, pk=None):
        nf = self.get_object()
        return _gerar_pdf_nf(nf)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    from datetime import timedelta
    oficina = get_oficina(request)
    hoje = timezone.now()
    inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0)
    inicio_mes_passado = (inicio_mes - timedelta(days=1)).replace(day=1)

    ordens_mes = OrdemServico.objects.filter(oficina=oficina, data_entrada__gte=inicio_mes)
    ordens_mes_passado = OrdemServico.objects.filter(
        oficina=oficina, data_entrada__gte=inicio_mes_passado, data_entrada__lt=inicio_mes
    )

    def calc_fat(ordens_qs):
        from django.db.models import ExpressionWrapper
        dec = DecimalField(max_digits=12, decimal_places=2)
        result = ordens_qs.aggregate(
            total_servicos=Coalesce(
                Sum(ExpressionWrapper(F('servicos__quantidade') * F('servicos__preco_unitario'), output_field=dec)),
                0, output_field=dec,
            ),
            total_pecas=Coalesce(
                Sum(ExpressionWrapper(F('pecas_usadas__quantidade') * F('pecas_usadas__preco_unitario'), output_field=dec)),
                0, output_field=dec,
            ),
            total_desconto=Coalesce(Sum('desconto', output_field=dec), 0, output_field=dec),
        )
        return float((result['total_servicos'] or 0) + (result['total_pecas'] or 0) - (result['total_desconto'] or 0))

    pecas_baixo = Peca.objects.filter(oficina=oficina, quantidade__lte=F('quantidade_minima')).count()

    # 1 query com annotate ao invés de 1 query por status
    status_counts = dict(
        OrdemServico.objects.filter(oficina=oficina)
        .values('status')
        .annotate(total=Count('id'))
        .values_list('status', 'total')
    )
    por_status = {s[0]: status_counts.get(s[0], 0) for s in OrdemServico.STATUS_CHOICES}

    faturamento_mensal = []
    for i in range(5, -1, -1):
        mes_inicio = (hoje - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0)
        mes_fim = (hoje - timedelta(days=30 * (i - 1))).replace(day=1, hour=0, minute=0, second=0) if i > 0 else hoje
        os_mes = OrdemServico.objects.filter(
            oficina=oficina, data_entrada__gte=mes_inicio, data_entrada__lt=mes_fim, status='concluida'
        )
        faturamento_mensal.append({
            'mes': mes_inicio.strftime('%b/%y'),
            'faturamento': calc_fat(os_mes),
            'ordens': os_mes.count(),
        })

    return JsonResponse({
        'resumo': {
            'total_clientes': Cliente.objects.filter(oficina=oficina).count(),
            'total_veiculos': Veiculo.objects.filter(oficina=oficina).count(),
            'ordens_abertas': OrdemServico.objects.filter(oficina=oficina, status__in=['aberta', 'em_andamento', 'aguardando_peca']).count(),
            'ordens_concluidas_mes': ordens_mes.filter(status='concluida').count(),
            'faturamento_mes': calc_fat(ordens_mes),
            'faturamento_mes_passado': calc_fat(ordens_mes_passado),
            'pecas_estoque_baixo': pecas_baixo,
            'total_funcionarios': Funcionario.objects.filter(oficina=oficina, ativo=True).count(),
        },
        'ordens_por_status': por_status,
        'faturamento_mensal': faturamento_mensal,
        'ultimas_ordens': OrdemServicoListSerializer(
            OrdemServico.objects.filter(oficina=oficina).select_related('cliente', 'veiculo', 'mecanico').order_by('-criado_em')[:5],
            many=True
        ).data,
        'agendamentos_hoje': [
            {
                'id': a.id,
                'hora': a.data_hora.strftime('%H:%M'),
                'cliente_nome': a.cliente.nome if a.cliente else a.nome_cliente,
                'veiculo_placa': a.veiculo_placa,
                'veiculo_info': a.veiculo_descricao or a.veiculo_placa,
                'servico': a.servico_desejado,
                'status': a.status,
                'status_display': a.get_status_display(),
            }
            for a in Agendamento.objects.filter(
                oficina=oficina,
                data_hora__date=hoje.date()
            ).select_related('cliente').order_by('data_hora')
        ],
    })


# ─── PDF helpers ──────────────────────────────────────────────────────────────

def _gerar_pdf_os(ordem):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    titulo_style = ParagraphStyle('titulo', parent=styles['Title'], fontSize=16, spaceAfter=6)
    subtitulo_style = ParagraphStyle('subtitulo', parent=styles['Heading2'], fontSize=11, spaceAfter=4)
    normal_style = styles['Normal']

    oficina_obj = ordem.oficina
    cor_hex = getattr(oficina_obj, 'cor_primaria', '#2563eb') or '#2563eb'
    cor_primaria = colors.HexColor(cor_hex)

    logo_elem = None
    if oficina_obj.logo:
        try:
            import os; from django.conf import settings
            logo_path = os.path.join(settings.MEDIA_ROOT, str(oficina_obj.logo))
            if os.path.exists(logo_path):
                from reportlab.platypus import Image as RLImage
                logo_elem = RLImage(logo_path, width=3*cm, height=2.5*cm, kind='proportional')
        except Exception:
            pass

    info_oficina = Paragraph(
        f'<b>{oficina_obj.nome}</b><br/>'
        f'<font size=8 color="#475569">CNPJ: {oficina_obj.cnpj or "-"}  |  Tel: {oficina_obj.telefone or "-"}<br/>'
        f'{oficina_obj.cidade or ""}{(" / "+oficina_obj.estado) if oficina_obj.estado else ""}</font>',
        ParagraphStyle('cab', parent=styles['Normal'], fontSize=11, leading=14)
    )
    titulo_os = Paragraph(
        f'<b>ORDEM DE SERVIÇO</b><br/><font size=10>Nº {ordem.numero}</font>',
        ParagraphStyle('osl', parent=styles['Normal'], fontSize=12, alignment=2, leading=16, textColor=cor_primaria)
    )

    if logo_elem:
        cab_data = [[logo_elem, info_oficina, titulo_os]]
        col_widths = [3*cm, 10*cm, 6*cm]
    else:
        cab_data = [[info_oficina, titulo_os]]
        col_widths = [12*cm, 7*cm]

    t_cab = Table(cab_data, colWidths=col_widths)
    t_cab.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, cor_primaria),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t_cab)
    story.append(Spacer(1, 0.4*cm))

    info_data = [
        ['Cliente:', ordem.cliente.nome, 'Data entrada:', ordem.data_entrada.strftime('%d/%m/%Y %H:%M')],
        ['Telefone:', ordem.cliente.celular or ordem.cliente.telefone, 'Status:', ordem.get_status_display()],
        ['Veículo:', f'{ordem.veiculo.marca} {ordem.veiculo.modelo}', 'Placa:', ordem.veiculo.placa],
        ['Mecânico:', ordem.mecanico.nome if ordem.mecanico else '-', 'KM entrada:', str(ordem.quilometragem_entrada)],
    ]
    t = Table(info_data, colWidths=[3.5*cm, 7*cm, 3.5*cm, 5*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph('Problema Relatado:', subtitulo_style))
    story.append(Paragraph(ordem.problema_relatado, normal_style))
    story.append(Spacer(1, 0.3*cm))

    servicos = ordem.servicos.all()
    if servicos:
        story.append(Paragraph('Serviços:', subtitulo_style))
        srv_data = [['Descrição', 'Qtd', 'Valor Unit.', 'Total']]
        for s in servicos:
            srv_data.append([s.descricao, str(s.quantidade), f'R$ {s.preco_unitario:.2f}', f'R$ {s.total:.2f}'])
        t = Table(srv_data, colWidths=[10*cm, 2*cm, 3.5*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    pecas = ordem.pecas_usadas.all()
    if pecas:
        story.append(Paragraph('Peças:', subtitulo_style))
        peca_data = [['Descrição', 'Qtd', 'Valor Unit.', 'Total']]
        for p in pecas:
            peca_data.append([p.descricao, str(p.quantidade), f'R$ {p.preco_unitario:.2f}', f'R$ {p.total:.2f}'])
        t = Table(peca_data, colWidths=[10*cm, 2*cm, 3.5*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a34a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    total_data = []
    if ordem.total_servicos > 0:
        total_data.append(['', 'Total Serviços:', f'R$ {ordem.total_servicos:.2f}'])
    if ordem.total_pecas > 0:
        total_data.append(['', 'Total Peças:', f'R$ {ordem.total_pecas:.2f}'])
    if ordem.desconto > 0:
        total_data.append(['', 'Desconto:', f'- R$ {ordem.desconto:.2f}'])
    total_data.append(['', 'TOTAL GERAL:', f'R$ {ordem.total_geral:.2f}'])

    t = Table(total_data, colWidths=[10*cm, 5*cm, 4*cm])
    t.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (1, -1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, -1), (2, -1), 12),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LINEABOVE', (1, -1), (2, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="OS_{ordem.numero}.pdf"'
    return response


def _gerar_pdf_nf(nf):
    ordem = nf.ordem
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    subtitulo_style = ParagraphStyle('subtitulo', parent=styles['Heading2'], fontSize=11, spaceAfter=4)
    normal_style = styles['Normal']

    oficina_obj = ordem.oficina
    cor_hex = getattr(oficina_obj, 'cor_primaria', '#2563eb') or '#2563eb'
    cor_primaria = colors.HexColor(cor_hex)

    logo_elem = None
    if oficina_obj.logo:
        try:
            import os; from django.conf import settings
            logo_path = os.path.join(settings.MEDIA_ROOT, str(oficina_obj.logo))
            if os.path.exists(logo_path):
                from reportlab.platypus import Image as RLImage
                logo_elem = RLImage(logo_path, width=3*cm, height=2.5*cm, kind='proportional')
        except Exception:
            pass

    info_oficina = Paragraph(
        f'<b>{oficina_obj.nome}</b><br/>'
        f'<font size=8 color="#475569">CNPJ: {oficina_obj.cnpj or "-"}  |  Tel: {oficina_obj.telefone or "-"}<br/>'
        f'{oficina_obj.cidade or ""}{(" / "+oficina_obj.estado) if oficina_obj.estado else ""}</font>',
        ParagraphStyle('cab', parent=styles['Normal'], fontSize=11, leading=14)
    )
    titulo_nf = Paragraph(
        f'<b>NOTA FISCAL DE SERVIÇO</b><br/>'
        f'<font size=10>Nº {nf.numero_nota}</font><br/>'
        f'<font size=8>Emissão: {nf.data_emissao.strftime("%d/%m/%Y %H:%M")}</font>',
        ParagraphStyle('nfl', parent=styles['Normal'], fontSize=12, alignment=2, leading=16, textColor=cor_primaria)
    )

    if logo_elem:
        cab_data = [[logo_elem, info_oficina, titulo_nf]]
        col_widths = [3*cm, 10*cm, 6*cm]
    else:
        cab_data = [[info_oficina, titulo_nf]]
        col_widths = [12*cm, 7*cm]

    t_cab = Table(cab_data, colWidths=col_widths)
    t_cab.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, cor_primaria),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t_cab)
    story.append(Spacer(1, 0.3*cm))

    info_data = [
        ['Oficina:', ordem.oficina.nome],
        ['CNPJ:', ordem.oficina.cnpj],
        ['Cliente:', ordem.cliente.nome],
        ['CPF/CNPJ:', ordem.cliente.cpf_cnpj or '-'],
        ['Veículo:', f'{ordem.veiculo.marca} {ordem.veiculo.modelo} - {ordem.veiculo.placa}'],
        ['OS Referência:', ordem.numero],
    ]
    t = Table(info_data, colWidths=[4*cm, 15*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    servicos = ordem.servicos.all()
    if servicos:
        story.append(Paragraph('Serviços Prestados:', subtitulo_style))
        srv_data = [['Descrição', 'Qtd', 'Valor Unit.', 'Total']]
        for s in servicos:
            srv_data.append([s.descricao, str(s.quantidade), f'R$ {s.preco_unitario:.2f}', f'R$ {s.total:.2f}'])
        t = Table(srv_data, colWidths=[10*cm, 2*cm, 3.5*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    pecas = ordem.pecas_usadas.all()
    if pecas:
        story.append(Paragraph('Peças Utilizadas:', subtitulo_style))
        peca_data = [['Descrição', 'Qtd', 'Valor Unit.', 'Total']]
        for p in pecas:
            peca_data.append([p.descricao, str(p.quantidade), f'R$ {p.preco_unitario:.2f}', f'R$ {p.total:.2f}'])
        t = Table(peca_data, colWidths=[10*cm, 2*cm, 3.5*cm, 3.5*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#15803d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    total_data = []
    if ordem.total_servicos > 0:
        total_data.append(['', 'Total Serviços:', f'R$ {ordem.total_servicos:.2f}'])
    if ordem.total_pecas > 0:
        total_data.append(['', 'Total Peças:', f'R$ {ordem.total_pecas:.2f}'])
    if ordem.desconto > 0:
        total_data.append(['', 'Desconto:', f'- R$ {ordem.desconto:.2f}'])
    total_data.append(['', 'TOTAL GERAL:', f'R$ {ordem.total_geral:.2f}'])

    t = Table(total_data, colWidths=[10*cm, 5*cm, 4*cm])
    t.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (1, -1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, -1), (2, -1), 13),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LINEABOVE', (1, -1), (2, -1), 1.5, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="NF_{nf.numero_nota}.pdf"'
    return response


# ─── Relatório de Saúde do Veículo ───────────────────────────────────────────

# Mapeamento: palavra-chave no nome do serviço/peça → sugestão futura
_SUGESTOES_MAPA = [
    (['óleo', 'oleo', 'troca de óleo', 'lubrificante'], 'Próxima troca de óleo: {km_prox} km ({km_restante} km restantes)'),
    (['correia dentada', 'correia do motor'], 'Próxima troca de correia dentada: {km_prox} km ({km_restante} km restantes)'),
    (['correia acessórios', 'correia alternador', 'correia poly-v'], 'Verificar correia de acessórios a cada 60.000 km'),
    (['filtro de ar', 'filtro ar'], 'Próxima troca de filtro de ar: {km_prox} km ({km_restante} km restantes)'),
    (['filtro combustível', 'filtro de combustível', 'filtro gasolina'], 'Próxima troca de filtro de combustível: {km_prox} km ({km_restante} km restantes)'),
    (['filtro de óleo', 'filtro oleo'], 'Próxima troca de filtro de óleo: {km_prox} km ({km_restante} km restantes)'),
    (['pastilha de freio', 'pastilha freio', 'pastilhas'], 'Verificar desgaste das pastilhas de freio nas próximas revisões'),
    (['disco de freio', 'disco freio', 'discos'], 'Verificar discos de freio na próxima revisão'),
    (['amortecedor', 'suspensão', 'mola'], 'Verificar amortecedores e suspensão a cada 40.000 km'),
    (['vela', 'velas de ignição', 'ignição'], 'Próxima troca de velas: {km_prox} km ({km_restante} km restantes)'),
    (['bateria'], 'Verificar estado da bateria a cada 2 anos ou 40.000 km'),
    (['alinhamento', 'balanceamento'], 'Realizar alinhamento e balanceamento a cada 10.000 km ou 6 meses'),
    (['pneu', 'pneus'], 'Verificar calibragem e desgaste dos pneus a cada 5.000 km'),
    (['fluido de freio', 'fluido freio'], 'Trocar fluido de freio a cada 2 anos ou 40.000 km'),
    (['líquido de arrefecimento', 'radiador', 'arrefecimento'], 'Trocar líquido de arrefecimento a cada 2 anos ou 40.000 km'),
    (['embreagem'], 'Verificar embreagem a cada 80.000 km'),
    (['revisão', 'revisao geral'], 'Próxima revisão geral: {km_prox} km ({km_restante} km restantes)'),
]

_INTERVALOS_KM = {
    'óleo': 10_000,
    'oleo': 10_000,
    'troca de óleo': 10_000,
    'filtro de ar': 15_000,
    'filtro ar': 15_000,
    'filtro combustível': 30_000,
    'filtro de combustível': 30_000,
    'filtro de óleo': 10_000,
    'filtro oleo': 10_000,
    'correia dentada': 60_000,
    'correia do motor': 60_000,
    'vela': 30_000,
    'velas de ignição': 30_000,
    'revisão': 10_000,
    'revisao geral': 10_000,
}


def _detectar_sugestoes(ordens, km_atual):
    """Analisa serviços/peças já feitos e monta lista de sugestões futuras."""
    feitos = set()
    km_por_servico = {}

    for os in ordens:
        km_os = os.quilometragem_entrada or 0
        for srv in os.servicos.all():
            texto = srv.descricao.lower()
            feitos.add(texto)
            km_por_servico[texto] = max(km_por_servico.get(texto, 0), km_os)
        for peca in os.pecas_usadas.all():
            texto = peca.descricao.lower()
            feitos.add(texto)
            km_por_servico[texto] = max(km_por_servico.get(texto, 0), km_os)

    sugestoes = []
    vistas = set()

    for palavras_chave, template in _SUGESTOES_MAPA:
        for chave in palavras_chave:
            if chave in vistas:
                continue
            # Verifica se algum item feito contém essa palavra-chave
            correspondeu = None
            km_feito = 0
            for texto_feito, km_f in km_por_servico.items():
                if chave in texto_feito:
                    correspondeu = texto_feito
                    km_feito = km_f
                    break

            if correspondeu:
                vistas.add(chave)
                intervalo = _INTERVALOS_KM.get(chave)
                if intervalo and '{km_prox}' in template:
                    km_prox = km_feito + intervalo
                    km_restante = km_prox - km_atual
                    if km_restante > 0:
                        sugestao = template.format(
                            km_prox=f'{km_prox:,}'.replace(',', '.'),
                            km_restante=f'{km_restante:,}'.replace(',', '.'),
                        )
                    else:
                        sugestao = f'⚠ VENCIDO — {template.split(":")[0]}: deveria ter sido feito há {abs(km_restante):,} km'.replace(',', '.')
                else:
                    sugestao = template.split(' ({km_restante}')[0].replace(' {km_prox} km', '')
                sugestoes.append(sugestao)

    return sugestoes


def _gerar_pdf_saude(veiculo):
    from decimal import Decimal

    ordens = OrdemServico.objects.filter(veiculo=veiculo).prefetch_related(
        'servicos', 'pecas_usadas'
    ).order_by('data_entrada')

    km_atual = veiculo.quilometragem

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()

    # Paleta de cores
    AZUL_ESCURO  = colors.HexColor('#1e3a5f')
    AZUL_MEDIO   = colors.HexColor('#2563eb')
    AZUL_CLARO   = colors.HexColor('#dbeafe')
    VERDE        = colors.HexColor('#16a34a')
    VERDE_CLARO  = colors.HexColor('#dcfce7')
    AMARELO      = colors.HexColor('#d97706')
    AMARELO_CLARO = colors.HexColor('#fef9c3')
    VERMELHO     = colors.HexColor('#dc2626')
    CINZA_CLARO  = colors.HexColor('#f1f5f9')
    CINZA_BORDA  = colors.HexColor('#cbd5e1')

    titulo_style   = ParagraphStyle('t', parent=styles['Title'],  fontSize=20, textColor=AZUL_ESCURO, spaceAfter=2, leading=24)
    subtit_style   = ParagraphStyle('s', parent=styles['Heading2'], fontSize=13, textColor=AZUL_ESCURO, spaceAfter=4, spaceBefore=10)
    label_style    = ParagraphStyle('l', parent=styles['Normal'],  fontSize=8,  textColor=colors.HexColor('#64748b'))
    valor_style    = ParagraphStyle('v', parent=styles['Normal'],  fontSize=10, textColor=colors.black, fontName='Helvetica-Bold')
    normal_style   = ParagraphStyle('n', parent=styles['Normal'],  fontSize=9,  leading=13)
    pequeno_style  = ParagraphStyle('p', parent=styles['Normal'],  fontSize=8,  textColor=colors.HexColor('#475569'), leading=12)
    rodape_style   = ParagraphStyle('r', parent=styles['Normal'],  fontSize=7,  textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)

    story = []
    W = 19*cm  # largura útil

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    oficina = veiculo.oficina
    header_data = [[
        Paragraph(f'<b>{oficina.nome}</b>', ParagraphStyle('of', parent=styles['Normal'], fontSize=10, textColor=colors.white)),
        Paragraph('RELATÓRIO DE SAÚDE DO VEÍCULO', ParagraphStyle('rh', parent=styles['Normal'], fontSize=14, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph(f'Gerado em<br/><b>{timezone.localtime().strftime("%d/%m/%Y %H:%M")}</b>', ParagraphStyle('dt', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#bfdbfe'), alignment=2)),
    ]]
    t_header = Table(header_data, colWidths=[5*cm, 9*cm, 5*cm])
    t_header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), AZUL_ESCURO),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',    (0, 0), (-1, -1), 10),
        ('ROUNDEDCORNERS', [6]),
    ]))
    story.append(t_header)
    story.append(Spacer(1, 0.5*cm))

    # ── Card do veículo ────────────────────────────────────────────────────────
    tipo_icone = {'moto': '🏍', 'carro': '🚗', 'caminhao': '🚛', 'outro': '🚙'}.get(veiculo.tipo, '🚗')
    veiculo_info = [
        [
            Paragraph(f'{tipo_icone} <b>{veiculo.marca} {veiculo.modelo}</b>', ParagraphStyle('vi', parent=styles['Normal'], fontSize=14, textColor=AZUL_ESCURO, fontName='Helvetica-Bold')),
            '', '', '',
        ],
        [
            Table([
                [Paragraph('PLACA', label_style)],
                [Paragraph(veiculo.placa.upper(), valor_style)],
            ], colWidths=[4*cm]),
            Table([
                [Paragraph('ANO', label_style)],
                [Paragraph(str(veiculo.ano or '—'), valor_style)],
            ], colWidths=[3*cm]),
            Table([
                [Paragraph('COR', label_style)],
                [Paragraph(veiculo.cor or '—', valor_style)],
            ], colWidths=[4*cm]),
            Table([
                [Paragraph('KM ATUAL', label_style)],
                [Paragraph(f'{km_atual:,} km'.replace(',', '.'), valor_style)],
            ], colWidths=[4*cm]),
        ],
        [
            Paragraph(f'Cliente: <b>{veiculo.cliente.nome}</b>', pequeno_style),
            Paragraph(f'Tel: {veiculo.cliente.celular or veiculo.cliente.telefone or "—"}', pequeno_style),
            Paragraph(f'Chassi: {veiculo.chassi or "—"}', pequeno_style),
            Paragraph(f'Total de OS: <b>{ordens.count()}</b>', pequeno_style),
        ],
    ]
    t_veiculo = Table(veiculo_info, colWidths=[4*cm, 3*cm, 4*cm, 4*cm])
    t_veiculo.setStyle(TableStyle([
        ('SPAN',       (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, -1), AZUL_CLARO),
        ('PADDING',    (0, 0), (-1, -1), 8),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('LINEBELOW',  (0, 0), (-1, 0), 0.5, CINZA_BORDA),
        ('LINEBELOW',  (0, 1), (-1, 1), 0.5, CINZA_BORDA),
        ('ROUNDEDCORNERS', [6]),
    ]))
    story.append(t_veiculo)
    story.append(Spacer(1, 0.5*cm))

    # ── Resumo financeiro ──────────────────────────────────────────────────────
    total_gasto = sum(o.total_geral for o in ordens)
    total_servicos = sum(o.total_servicos for o in ordens)
    total_pecas    = sum(o.total_pecas    for o in ordens)
    n_concluidas   = ordens.filter(status='concluida').count()

    resumo_data = [
        [
            Table([[Paragraph('TOTAL INVESTIDO', label_style)], [Paragraph(f'R$ {total_gasto:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'), ParagraphStyle('rv', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', textColor=VERDE))]], colWidths=[4.5*cm]),
            Table([[Paragraph('EM SERVIÇOS',     label_style)], [Paragraph(f'R$ {total_servicos:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'), valor_style)]], colWidths=[4.5*cm]),
            Table([[Paragraph('EM PEÇAS',        label_style)], [Paragraph(f'R$ {total_pecas:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),    valor_style)]], colWidths=[4.5*cm]),
            Table([[Paragraph('OS CONCLUÍDAS',   label_style)], [Paragraph(str(n_concluidas), valor_style)]], colWidths=[4.5*cm]),
        ]
    ]
    t_resumo = Table(resumo_data, colWidths=[4.75*cm, 4.75*cm, 4.75*cm, 4.75*cm])
    t_resumo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), CINZA_CLARO),
        ('PADDING',    (0, 0), (-1, -1), 10),
        ('LINEBEFORE', (1, 0), (3, 0), 0.5, CINZA_BORDA),
        ('ROUNDEDCORNERS', [6]),
    ]))
    story.append(t_resumo)
    story.append(Spacer(1, 0.5*cm))

    # ── Histórico de OS ────────────────────────────────────────────────────────
    if ordens.exists():
        story.append(Paragraph('Histórico de Atendimentos', subtit_style))

        STATUS_COR = {
            'concluida':        (VERDE,       VERDE_CLARO),
            'cancelada':        (VERMELHO,     colors.HexColor('#fee2e2')),
            'aberta':           (AZUL_MEDIO,   AZUL_CLARO),
            'em_andamento':     (AMARELO,      AMARELO_CLARO),
            'aguardando_peca':  (AMARELO,      AMARELO_CLARO),
        }

        for os in ordens.order_by('-data_entrada'):
            cor_txt, cor_bg = STATUS_COR.get(os.status, (colors.grey, CINZA_CLARO))

            # Linha de cabeçalho da OS
            srv_resumo = ', '.join(s.descricao for s in os.servicos.all()[:3])
            if os.servicos.count() > 3:
                srv_resumo += f' +{os.servicos.count()-3}'
            pec_resumo = ', '.join(p.descricao for p in os.pecas_usadas.all()[:3])
            if os.pecas_usadas.count() > 3:
                pec_resumo += f' +{os.pecas_usadas.count()-3}'

            os_header = [
                [
                    Paragraph(f'<b>OS {os.numero}</b>', ParagraphStyle('oh', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=AZUL_ESCURO)),
                    Paragraph(os.data_entrada.strftime('%d/%m/%Y'), pequeno_style),
                    Paragraph(f'KM: {os.quilometragem_entrada:,}'.replace(',', '.'), pequeno_style),
                    Paragraph(f'<b>{os.get_status_display()}</b>', ParagraphStyle('st', parent=styles['Normal'], fontSize=8, textColor=cor_txt, fontName='Helvetica-Bold')),
                    Paragraph(f'<b>R$ {os.total_geral:,.2f}</b>'.replace(',', 'X').replace('.', ',').replace('X', '.'), ParagraphStyle('tg', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=VERDE)),
                ],
            ]
            t_os_h = Table(os_header, colWidths=[3.5*cm, 2.5*cm, 3*cm, 3.5*cm, 3.5*cm])
            t_os_h.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), cor_bg),
                ('PADDING',    (0, 0), (-1, -1), 6),
                ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
                ('LINEBELOW',  (0, 0), (-1, -1), 0.3, CINZA_BORDA),
            ]))
            story.append(t_os_h)

            # Detalhe: serviços e peças
            det_rows = []
            if os.problema_relatado:
                det_rows.append([Paragraph(f'Problema: {os.problema_relatado}', pequeno_style), ''])
            if srv_resumo:
                det_rows.append([Paragraph(f'Serviços: {srv_resumo}', pequeno_style), ''])
            if pec_resumo:
                det_rows.append([Paragraph(f'Peças: {pec_resumo}', pequeno_style), ''])
            if os.mecanico:
                det_rows.append([Paragraph(f'Mecânico: {os.mecanico.nome}', pequeno_style), ''])

            if det_rows:
                t_det = Table(det_rows, colWidths=[15*cm, 4*cm])
                t_det.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('PADDING',    (0, 0), (-1, -1), 5),
                    ('SPAN',       (0, 0), (1, 0)),
                ]))
                story.append(t_det)

            story.append(Spacer(1, 0.2*cm))

    # ── Sugestões futuras ──────────────────────────────────────────────────────
    sugestoes = _detectar_sugestoes(ordens, km_atual)

    story.append(Paragraph('Sugestões e Próximas Revisões', subtit_style))

    if sugestoes:
        sug_rows = []
        for i, s in enumerate(sugestoes):
            icone = '⚠' if s.startswith('⚠') else '✓'
            cor_linha = colors.HexColor('#fef2f2') if icone == '⚠' else AMARELO_CLARO
            sug_rows.append([
                Paragraph(icone, ParagraphStyle('ic', parent=styles['Normal'], fontSize=10, textColor=AMARELO if icone == '✓' else VERMELHO)),
                Paragraph(s.lstrip('⚠ '), pequeno_style),
            ])
        t_sug = Table(sug_rows, colWidths=[0.8*cm, 18.2*cm])
        t_sug.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), AMARELO_CLARO),
            ('PADDING',    (0, 0), (-1, -1), 5),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW',  (0, 0), (-1, -2), 0.3, colors.HexColor('#fde68a')),
            ('ROUNDEDCORNERS', [4]),
        ]))
        story.append(t_sug)
    else:
        story.append(Paragraph(
            'Nenhum dado de histórico suficiente para gerar sugestões automáticas. '
            'Registre as ordens de serviço com descrições detalhadas para que o sistema '
            'possa recomendar revisões futuras.',
            normal_style
        ))

    # ── Aviso legal ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph(
        'Este relatório foi gerado automaticamente com base no histórico de ordens de serviço registradas no sistema DoMecânico. '
        'As sugestões de revisão são estimativas baseadas em intervalos médios e podem variar conforme o fabricante e condições de uso do veículo.',
        rodape_style
    ))
    story.append(Paragraph(f'DoMecânico — {oficina.nome} · {oficina.cidade}/{oficina.estado}', rodape_style))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    placa = veiculo.placa.replace('-', '').upper()
    response['Content-Disposition'] = f'inline; filename="Saude_{placa}.pdf"'
    return response


# ─── Checklist de Entrada ─────────────────────────────────────────────────────

class ChecklistViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ChecklistEntradaSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = ChecklistEntrada.objects.filter(oficina=oficina).select_related('cliente', 'veiculo', 'ordem')
        ordem_id = self.request.query_params.get('ordem')
        status_f = self.request.query_params.get('status')
        if ordem_id:
            qs = qs.filter(ordem_id=ordem_id)
        if status_f:
            qs = qs.filter(status=status_f)
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))

    @action(detail=True, methods=['post'], url_path='adicionar-dano')
    def adicionar_dano(self, request, pk=None):
        checklist = self.get_object()
        data = {**request.data, 'checklist': checklist.id}
        serializer = DanoChecklistSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save(checklist=checklist)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], url_path='remover-dano/(?P<dano_id>[^/.]+)')
    def remover_dano(self, request, pk=None, dano_id=None):
        checklist = self.get_object()
        try:
            dano = DanoChecklist.objects.get(id=dano_id, checklist=checklist)
            if dano.foto:
                dano.foto.delete()
            dano.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except DanoChecklist.DoesNotExist:
            return Response({'erro': 'Dano não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='upload-foto-dano/(?P<dano_id>[^/.]+)')
    def upload_foto_dano(self, request, pk=None, dano_id=None):
        checklist = self.get_object()
        try:
            dano = DanoChecklist.objects.get(id=dano_id, checklist=checklist)
        except DanoChecklist.DoesNotExist:
            return Response({'erro': 'Dano não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        foto = request.FILES.get('foto')
        if not foto:
            return Response({'erro': 'Nenhuma foto enviada.'}, status=status.HTTP_400_BAD_REQUEST)
        MAX_SIZE = 10 * 1024 * 1024  # 10 MB
        if foto.size > MAX_SIZE:
            return Response({'erro': 'Arquivo muito grande. Limite: 10 MB.'}, status=status.HTTP_400_BAD_REQUEST)
        if not _validar_imagem(foto):
            return Response({'erro': 'Arquivo inválido. Envie uma imagem JPEG, PNG, WebP ou GIF.'}, status=status.HTTP_400_BAD_REQUEST)
        if dano.foto:
            dano.foto.delete()
        dano.foto = foto
        dano.save()
        return Response(DanoChecklistSerializer(dano, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='assinar')
    def assinar(self, request, pk=None):
        checklist = self.get_object()
        assinatura = request.data.get('assinatura', '')
        if not assinatura:
            return Response({'erro': 'Assinatura não enviada.'}, status=status.HTTP_400_BAD_REQUEST)
        checklist.assinatura = assinatura
        checklist.status = 'assinado'
        checklist.data_assinatura = timezone.now()
        checklist.save()
        return Response(ChecklistEntradaSerializer(checklist, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='pdf')
    def gerar_pdf(self, request, pk=None):
        checklist = self.get_object()
        return _gerar_pdf_checklist(checklist)


@api_view(['GET'])
@permission_classes([AllowAny])
def checklist_publico(request, token):
    """Rota pública — o cliente acessa via link para ver e assinar o checklist."""
    try:
        checklist = ChecklistEntrada.objects.select_related(
            'cliente', 'veiculo', 'oficina'
        ).prefetch_related('danos').get(token_publico=token)
    except ChecklistEntrada.DoesNotExist:
        return Response({'erro': 'Checklist não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    return Response(ChecklistPublicoSerializer(checklist, context={'request': request}).data)


@api_view(['POST'])
@permission_classes([AllowAny])
def checklist_assinar_publico(request, token):
    """Rota pública — cliente assina digitalmente."""
    try:
        checklist = ChecklistEntrada.objects.get(token_publico=token)
    except ChecklistEntrada.DoesNotExist:
        return Response({'erro': 'Checklist não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    if checklist.status == 'assinado':
        return Response({'erro': 'Checklist já foi assinado.'}, status=status.HTTP_400_BAD_REQUEST)
    assinatura = request.data.get('assinatura', '')
    if not assinatura:
        return Response({'erro': 'Assinatura não enviada.'}, status=status.HTTP_400_BAD_REQUEST)
    checklist.assinatura = assinatura
    checklist.status = 'assinado'
    checklist.data_assinatura = timezone.now()
    checklist.save()
    return Response({'sucesso': True, 'mensagem': 'Checklist assinado com sucesso!'})


def _gerar_pdf_checklist(checklist):
    from reportlab.platypus import Image as RLImage
    import base64

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    sub_style = ParagraphStyle('sub', parent=styles['Heading2'], fontSize=11, spaceAfter=4)
    normal = styles['Normal']
    bold = ParagraphStyle('bold', parent=normal, fontName='Helvetica-Bold')

    oficina_obj = checklist.oficina
    cor_hex = getattr(oficina_obj, 'cor_primaria', '#2563eb') or '#2563eb'
    cor_primaria = colors.HexColor(cor_hex)

    logo_elem = None
    if oficina_obj.logo:
        try:
            import os; from django.conf import settings
            logo_path = os.path.join(settings.MEDIA_ROOT, str(oficina_obj.logo))
            if os.path.exists(logo_path):
                logo_elem = RLImage(logo_path, width=3*cm, height=2.5*cm, kind='proportional')
        except Exception:
            pass

    info_oficina = Paragraph(
        f'<b>{oficina_obj.nome}</b><br/>'
        f'<font size=8 color="#475569">CNPJ: {oficina_obj.cnpj or "-"}  |  Tel: {oficina_obj.telefone or "-"}<br/>'
        f'{oficina_obj.cidade or ""}{(" / "+oficina_obj.estado) if oficina_obj.estado else ""}</font>',
        ParagraphStyle('cab', parent=styles['Normal'], fontSize=11, leading=14)
    )
    titulo_chk = Paragraph(
        '<b>CHECKLIST DE ENTRADA</b><br/><font size=9>de Veículo</font>',
        ParagraphStyle('chkl', parent=styles['Normal'], fontSize=12, alignment=2, leading=16, textColor=cor_primaria)
    )

    if logo_elem:
        cab_data = [[logo_elem, info_oficina, titulo_chk]]
        col_widths = [3*cm, 10*cm, 6*cm]
    else:
        cab_data = [[info_oficina, titulo_chk]]
        col_widths = [12*cm, 7*cm]

    t_cab = Table(cab_data, colWidths=col_widths)
    t_cab.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, cor_primaria),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t_cab)
    story.append(Spacer(1, 0.3*cm))

    info = [
        ['Cliente:', checklist.cliente.nome, 'Data:', checklist.criado_em.strftime('%d/%m/%Y %H:%M')],
        ['Veículo:', f'{checklist.veiculo.marca} {checklist.veiculo.modelo}', 'Placa:', checklist.veiculo.placa],
        ['Quilometragem:', f'{checklist.quilometragem} km', 'Combustível:', checklist.get_nivel_combustivel_display()],
        ['Status:', checklist.get_status_display(), '', ''],
    ]
    t = Table(info, colWidths=[3.5*cm, 7.5*cm, 3*cm, 5*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    if checklist.observacoes_gerais:
        story.append(Paragraph('Observações Gerais:', sub_style))
        story.append(Paragraph(checklist.observacoes_gerais, normal))
        story.append(Spacer(1, 0.3*cm))

    danos = checklist.danos.all()
    if danos:
        story.append(Paragraph('Danos / Avarias Identificados:', sub_style))
        dano_data = [['Região', 'Tipo', 'Descrição']]
        for d in danos:
            dano_data.append([d.get_regiao_display(), d.get_tipo_display(), d.descricao or '-'])
        t = Table(dano_data, colWidths=[4*cm, 4*cm, 11*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')]),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph('Assinatura do Cliente:', sub_style))

    if checklist.assinatura and checklist.assinatura.startswith('data:image'):
        try:
            header, b64data = checklist.assinatura.split(',', 1)
            img_data = base64.b64decode(b64data)
            img_buf = io.BytesIO(img_data)
            img = RLImage(img_buf, width=8*cm, height=3*cm)
            story.append(img)
            if checklist.data_assinatura:
                story.append(Paragraph(
                    f'Assinado em: {checklist.data_assinatura.strftime("%d/%m/%Y %H:%M")}',
                    ParagraphStyle('data_ass', parent=normal, fontSize=8, textColor=colors.grey)
                ))
        except Exception:
            story.append(Paragraph('[Assinatura registrada digitalmente]', normal))
    else:
        story.append(Spacer(1, 2*cm))
        story.append(Table([['_' * 50]], colWidths=[10*cm]))
        story.append(Paragraph(checklist.cliente.nome, ParagraphStyle('nome_ass', parent=normal, fontSize=9)))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="checklist_{checklist.id}.pdf"'
    return response


# ─── Portal público de acompanhamento de OS ───────────────────────────────────

@api_view(['GET'])
@permission_classes([AllowAny])
def os_publica_por_token(request, token):
    """Retorna os dados públicos de uma OS via token único."""
    try:
        os_obj = OrdemServico.objects.select_related(
            'cliente', 'veiculo', 'mecanico', 'oficina'
        ).prefetch_related(
            'veiculo__fotos', 'servicos', 'servicos__garantia',
            'checklist', 'checklist__danos', 'orcamento_origem', 'orcamento_origem__itens',
        ).get(token_publico=token)
    except OrdemServico.DoesNotExist:
        return Response({'erro': 'OS não encontrada ou link inválido.'}, status=404)

    return Response(OSPublicaSerializer(os_obj, context={'request': request}).data)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([OSPublicaBuscarThrottle])
def os_publica_buscar(request):
    """Busca OS pelo número (ou placa) + CPF/CNPJ do cliente."""
    import re
    placa_raw = request.data.get('placa', '').strip().upper()
    placa_norm = re.sub(r'[^A-Z0-9]', '', placa_raw)

    cpf_cnpj_raw = request.data.get('cpf_cnpj', '').strip()
    cpf_cnpj_norm = re.sub(r'\D', '', cpf_cnpj_raw)

    if not placa_norm or not cpf_cnpj_norm:
        return Response({'erro': 'Informe a placa e o CPF/CNPJ.'}, status=400)

    # Filter by placa at DB level (placa stored without special chars), then verify CPF in Python
    candidatos = OrdemServico.objects.select_related(
        'cliente', 'veiculo', 'mecanico', 'oficina'
    ).prefetch_related(
        'veiculo__fotos', 'servicos', 'servicos__garantia',
        'checklist', 'checklist__danos', 'orcamento_origem', 'orcamento_origem__itens',
    ).filter(
        veiculo__placa__iexact=placa_norm,
    ).order_by('-criado_em')

    ordens = [
        o for o in candidatos
        if re.sub(r'\D', '', o.cliente.cpf_cnpj or '') == cpf_cnpj_norm
    ]

    if not ordens:
        return Response({'erro': 'Nenhuma OS encontrada com esses dados. Verifique a placa e o CPF/CNPJ.'}, status=404)

    return Response(OSPublicaSerializer(ordens, many=True, context={'request': request}).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def os_publica_por_placa_oficina(request, slug):
    """Busca OS por placa e/ou CPF dentro de uma oficina específica (mini-site público)."""
    import re
    from accounts.models import Oficina
    from django.db.models import Q

    placa_raw = request.query_params.get('placa', '').strip().upper()
    cpf_raw = request.query_params.get('cpf', '').strip()
    placa_norm = re.sub(r'[^A-Z0-9]', '', placa_raw)
    cpf_norm = re.sub(r'\D', '', cpf_raw)

    if not placa_norm and not cpf_norm:
        return Response({'erro': 'Informe a placa ou o CPF.'}, status=400)

    try:
        oficina = Oficina.objects.get(slug_publico=slug, perfil_publico_ativo=True)
    except Oficina.DoesNotExist:
        return Response({'erro': 'Oficina não encontrada.'}, status=404)

    qs = OrdemServico.objects.select_related('veiculo', 'cliente').filter(oficina=oficina)

    filtro = Q()
    if placa_norm:
        filtro |= Q(veiculo__placa__iexact=placa_norm)
    if cpf_norm:
        filtro |= Q(cliente__cpf_cnpj__icontains=cpf_norm)

    ordens = qs.filter(filtro).order_by('-criado_em')[:5]

    resultado = [
        {
            'id': o.id,
            'numero': o.numero,
            'status': o.status,
            'veiculo_placa': o.veiculo.placa,
            'veiculo_modelo': f'{o.veiculo.marca} {o.veiculo.modelo}',
            'descricao_problema': o.problema_relatado,
            'data_entrada': o.data_entrada,
            'token_publico': o.token_publico,
        }
        for o in ordens
    ]
    return Response(resultado)


# ── Agendamento ────────────────────────────────────────────────────────────────

class AgendamentoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = AgendamentoSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Agendamento.objects.filter(oficina=oficina)
        data = self.request.query_params.get('data')
        status_f = self.request.query_params.get('status')
        if data:
            qs = qs.filter(data_hora__date=data)
        if status_f:
            qs = qs.filter(status=status_f)
        return qs

    def perform_create(self, serializer):
        serializer.save(oficina=get_oficina(self.request))

    @action(detail=True, methods=['patch'])
    def confirmar(self, request, pk=None):
        ag = self.get_object()
        ag.status = 'confirmado'
        ag.save()
        try:
            from .whatsapp import notificar_agendamento
            notificar_agendamento(ag)
        except Exception:
            pass
        return Response(AgendamentoSerializer(ag).data)

    @action(detail=True, methods=['patch'])
    def cancelar(self, request, pk=None):
        ag = self.get_object()
        ag.status = 'cancelado'
        ag.save()
        return Response(AgendamentoSerializer(ag).data)


# ── Orçamento Digital ──────────────────────────────────────────────────────────

class OrcamentoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = OrcamentoSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = Orcamento.objects.filter(oficina=oficina).select_related('cliente', 'veiculo')
        search = self.request.query_params.get('search')
        status_f = self.request.query_params.get('status')
        if search:
            qs = qs.filter(Q(numero__icontains=search) | Q(cliente__nome__icontains=search))
        if status_f:
            qs = qs.filter(status=status_f)
        return qs

    def perform_create(self, serializer):
        oficina = get_oficina(self.request)
        ultimo = Orcamento.objects.filter(oficina=oficina).order_by('-id').values_list('id', flat=True).first() or 0
        numero = f'ORC{ultimo + 1:06d}'
        serializer.save(oficina=oficina, numero=numero)
        try:
            from .whatsapp import notificar_orcamento
            notificar_orcamento(serializer.instance)
        except Exception:
            pass

    @action(detail=True, methods=['post'])
    def adicionar_item(self, request, pk=None):
        orc = self.get_object()
        s = ItemOrcamentoSerializer(data={**request.data, 'orcamento': orc.id})
        s.is_valid(raise_exception=True)
        s.save(orcamento=orc)
        return Response(OrcamentoSerializer(orc, context={'request': request}).data)

    @action(detail=True, methods=['delete'], url_path='remover-item/(?P<item_id>[0-9]+)')
    def remover_item(self, request, pk=None, item_id=None):
        orc = self.get_object()
        ItemOrcamento.objects.filter(id=item_id, orcamento=orc).delete()
        return Response(OrcamentoSerializer(orc, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def aprovar(self, request, pk=None):
        orc = self.get_object()
        if orc.status not in ('pendente',):
            return Response({'erro': 'Só orçamentos pendentes podem ser aprovados.'}, status=400)
        orc.status = 'aprovado'
        orc.save()
        return Response(OrcamentoSerializer(orc, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def rejeitar(self, request, pk=None):
        orc = self.get_object()
        orc.status = 'rejeitado'
        orc.save()
        return Response(OrcamentoSerializer(orc, context={'request': request}).data)

    @action(detail=True, methods=['patch'])
    def atualizar_desconto(self, request, pk=None):
        orc = self.get_object()
        orc.desconto = request.data.get('desconto', 0)
        orc.save()
        return Response(OrcamentoSerializer(orc, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def converter_os(self, request, pk=None):
        orc = self.get_object()
        if orc.status != 'aprovado':
            return Response({'erro': 'Orçamento precisa estar aprovado.'}, status=400)
        if orc.ordem:
            return Response({'erro': 'Já convertido em OS.'}, status=400)
        oficina = get_oficina(request)
        ultimo = OrdemServico.objects.filter(oficina=oficina).order_by('-id').values_list('id', flat=True).first() or 0
        numero = f'OS{ultimo + 1:06d}'
        os_obj = OrdemServico.objects.create(
            oficina=oficina,
            numero=numero,
            cliente=orc.cliente,
            veiculo=orc.veiculo,
            mecanico=orc.mecanico,
            problema_relatado=orc.problema_relatado,
            observacoes=orc.observacoes,
            desconto=orc.desconto,
        )
        for item in orc.itens.all():
            if item.tipo == 'servico':
                ServicoOS.objects.create(
                    ordem=os_obj, descricao=item.descricao,
                    quantidade=item.quantidade, preco_unitario=item.preco_unitario,
                )
            else:
                PecaOS.objects.create(
                    ordem=os_obj, descricao=item.descricao,
                    quantidade=item.quantidade, preco_unitario=item.preco_unitario,
                )
        orc.ordem = os_obj
        orc.save()
        return Response({'os_id': os_obj.id, 'os_numero': os_obj.numero})


@api_view(['GET'])
@permission_classes([AllowAny])
def orcamento_publico(request, token):
    try:
        orc = Orcamento.objects.select_related(
            'cliente', 'veiculo', 'oficina'
        ).prefetch_related('itens').get(token_publico=token)
    except Orcamento.DoesNotExist:
        return Response({'erro': 'Orçamento não encontrado.'}, status=404)
    return Response(OrcamentoPublicoSerializer(orc, context={'request': request}).data)


@api_view(['POST'])
@permission_classes([AllowAny])
def orcamento_responder(request, token):
    try:
        orc = Orcamento.objects.get(token_publico=token, status='pendente')
    except Orcamento.DoesNotExist:
        return Response({'erro': 'Orçamento não encontrado ou já respondido.'}, status=404)
    resposta = request.data.get('resposta')
    if resposta not in ('aprovado', 'rejeitado'):
        return Response({'erro': 'Resposta inválida.'}, status=400)
    orc.status = resposta
    orc.save()
    return Response({'status': orc.status})


# ── Garantia ───────────────────────────────────────────────────────────────────

class GarantiaViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = GarantiaServicoSerializer
    http_method_names = ['get', 'post', 'delete', 'head', 'options']  # sem PUT/PATCH

    def get_queryset(self):
        oficina = get_oficina(self.request)
        return GarantiaServico.objects.filter(
            servico_os__ordem__oficina=oficina
        ).select_related('servico_os__ordem__cliente', 'servico_os__ordem__veiculo')


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def garantia_default(request):
    from .serializers import GarantiaDefaultSerializer
    oficina = get_oficina(request)
    config, _ = GarantiaDefault.objects.get_or_create(
        oficina=oficina,
        defaults={'prazo_dias': 90, 'observacoes': ''},
    )
    if request.method == 'GET':
        return Response(GarantiaDefaultSerializer(config).data)
    s = GarantiaDefaultSerializer(config, data=request.data, partial=True)
    if s.is_valid():
        s.save()
        return Response(s.data)
    return Response(s.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def aplicar_garantia_servico(request, servico_id):
    """Aplica garantia em um ServicoOS, usando prazo padrão da oficina se não informado."""
    oficina = get_oficina(request)
    try:
        servico = ServicoOS.objects.get(id=servico_id, ordem__oficina=oficina)
    except ServicoOS.DoesNotExist:
        return Response({'erro': 'Serviço não encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if hasattr(servico, 'garantia'):
        return Response({'erro': 'Este serviço já possui garantia.'}, status=status.HTTP_400_BAD_REQUEST)

    config, _ = GarantiaDefault.objects.get_or_create(
        oficina=oficina, defaults={'prazo_dias': 90}
    )
    prazo = request.data.get('prazo_dias', config.prazo_dias)
    obs = request.data.get('observacoes', config.observacoes)
    data_inicio = request.data.get('data_inicio', None)

    from datetime import date
    garantia = GarantiaServico.objects.create(
        servico_os=servico,
        prazo_dias=int(prazo),
        observacoes=obs,
        data_inicio=data_inicio or date.today(),
    )
    from .serializers import GarantiaServicoSerializer
    return Response(GarantiaServicoSerializer(garantia).data, status=status.HTTP_201_CREATED)


# ── Comissão ───────────────────────────────────────────────────────────────────

class ComissaoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ComissaoMecanicoSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        qs = ComissaoMecanico.objects.filter(
            ordem__oficina=oficina
        ).select_related('funcionario', 'ordem')
        funcionario_id = self.request.query_params.get('funcionario')
        pago = self.request.query_params.get('pago')
        mes = self.request.query_params.get('mes')
        data_inicio = self.request.query_params.get('data_inicio')
        data_fim = self.request.query_params.get('data_fim')
        if funcionario_id:
            qs = qs.filter(funcionario_id=funcionario_id)
        if pago is not None:
            qs = qs.filter(pago=pago == 'true')
        if mes:
            qs = qs.filter(criado_em__month=mes.split('-')[1], criado_em__year=mes.split('-')[0])
        if data_inicio:
            qs = qs.filter(criado_em__date__gte=data_inicio)
        if data_fim:
            qs = qs.filter(criado_em__date__lte=data_fim)
        return qs

    @action(detail=True, methods=['post'])
    def marcar_pago(self, request, pk=None):
        from django.utils import timezone
        c = self.get_object()
        c.pago = True
        c.data_pagamento = timezone.now().date()
        c.save()
        return Response(ComissaoMecanicoSerializer(c).data)

    @action(detail=True, methods=['delete'])
    def remover(self, request, pk=None):
        c = self.get_object()
        c.delete()
        return Response(status=204)

    @action(detail=False, methods=['post'])
    def calcular_os(self, request):
        """Calcula e registra comissão para uma OS. Suporta múltiplos mecânicos via lista."""
        os_id = request.data.get('os_id')
        # lista de { funcionario_id, percentual } ou campos simples para retrocompatibilidade
        participantes = request.data.get('participantes')  # lista nova
        if not os_id:
            return Response({'erro': 'os_id é obrigatório.'}, status=400)
        try:
            os_obj = OrdemServico.objects.get(id=os_id, oficina=get_oficina(request))
        except OrdemServico.DoesNotExist:
            return Response({'erro': 'OS não encontrada.'}, status=404)

        if not participantes:
            # retrocompatibilidade: campo simples com mecânico da OS
            funcionario_id = request.data.get('funcionario_id')
            percentual = request.data.get('percentual')
            if not percentual:
                return Response({'erro': 'percentual é obrigatório.'}, status=400)
            if funcionario_id:
                try:
                    func = Funcionario.objects.get(id=funcionario_id, oficina=get_oficina(request))
                except Funcionario.DoesNotExist:
                    return Response({'erro': 'Funcionário não encontrado.'}, status=404)
            elif os_obj.mecanico:
                func = os_obj.mecanico
            else:
                return Response({'erro': 'Informe o funcionário ou vincule um mecânico à OS.'}, status=400)
            valor = float(os_obj.total_servicos) * float(percentual) / 100
            c, _ = ComissaoMecanico.objects.update_or_create(
                funcionario=func, ordem=os_obj,
                defaults={'percentual': percentual, 'valor': round(valor, 2)},
            )
            return Response(ComissaoMecanicoSerializer(c).data)

        # múltiplos participantes
        criados = []
        erros = []
        for p in participantes:
            fid = p.get('funcionario_id')
            pct = p.get('percentual')
            if not fid or not pct:
                erros.append(f'Item inválido: {p}')
                continue
            try:
                func = Funcionario.objects.get(id=fid, oficina=get_oficina(request))
            except Funcionario.DoesNotExist:
                erros.append(f'Funcionário {fid} não encontrado.')
                continue
            valor = float(os_obj.total_servicos) * float(pct) / 100
            c, _ = ComissaoMecanico.objects.update_or_create(
                funcionario=func, ordem=os_obj,
                defaults={'percentual': pct, 'valor': round(valor, 2)},
            )
            criados.append(ComissaoMecanicoSerializer(c).data)
        return Response({'criados': criados, 'erros': erros})


# ── Alertas de Estoque ─────────────────────────────────────────────────────────

class AlertaEstoqueViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = AlertaEstoqueSerializer

    def get_queryset(self):
        oficina = get_oficina(self.request)
        apenas_nao_lidos = self.request.query_params.get('nao_lidos')
        qs = AlertaEstoque.objects.filter(
            peca__oficina=oficina
        ).select_related('peca')
        if apenas_nao_lidos == 'true':
            qs = qs.filter(lido=False)
        return qs

    @action(detail=True, methods=['post'])
    def marcar_lido(self, request, pk=None):
        alerta = self.get_object()
        alerta.lido = True
        alerta.save()
        return Response({'ok': True})

    @action(detail=False, methods=['post'])
    def marcar_todos_lidos(self, request):
        oficina = get_oficina(request)
        AlertaEstoque.objects.filter(peca__oficina=oficina, lido=False).update(lido=True)
        return Response({'ok': True})


# ── Histórico público por placa ────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def historico_por_placa(request, placa):
    import re
    placa_norm = re.sub(r'[^A-Z0-9]', '', placa.upper())
    oficina = get_oficina(request)
    resultado = list(OrdemServico.objects.select_related(
        'cliente', 'veiculo', 'mecanico', 'oficina'
    ).prefetch_related('servicos').filter(
        status='concluida',
        oficina=oficina,
        veiculo__placa__iexact=placa_norm,
    ).order_by('-data_conclusao'))
    if not resultado:
        return Response({'erro': 'Nenhum histórico encontrado para esta placa.'}, status=404)
    data = [{
        'oficina': o.oficina.nome,
        'numero': o.numero,
        'data_conclusao': o.data_conclusao,
        'quilometragem_entrada': o.quilometragem_entrada,
        'servicos': [{'descricao': s.descricao, 'quantidade': str(s.quantidade)} for s in o.servicos.all()],
        'observacoes': o.observacoes,
    } for o in resultado]
    veiculo = resultado[0].veiculo
    return Response({
        'placa': veiculo.placa,
        'veiculo_info': f'{veiculo.marca} {veiculo.modelo} {veiculo.ano or ""} — {veiculo.cor or ""}'.strip(),
        'historico': data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_dados(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    oficina = get_oficina(request)
    wb = openpyxl.Workbook()

    H_FONT = Font(bold=True, color='FFFFFF')
    H_FILL = PatternFill('solid', fgColor='2563EB')

    def cabecalho(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(horizontal='center')

    ws = wb.active
    ws.title = 'Clientes'
    cabecalho(ws, ['ID', 'Nome', 'CPF/CNPJ', 'Telefone', 'Celular', 'Email', 'Cidade', 'Estado'])
    for c in Cliente.objects.filter(oficina=oficina).order_by('nome'):
        ws.append([c.id, c.nome, c.cpf_cnpj, c.telefone, c.celular, c.email, c.cidade, c.estado])

    ws2 = wb.create_sheet('Veículos')
    cabecalho(ws2, ['ID', 'Placa', 'Marca', 'Modelo', 'Ano', 'Cor', 'Cliente'])
    for v in Veiculo.objects.filter(oficina=oficina).select_related('cliente').order_by('placa'):
        ws2.append([v.id, v.placa, v.marca, v.modelo, v.ano, v.cor, v.cliente.nome if v.cliente else ''])

    ws3 = wb.create_sheet('Ordens de Serviço')
    cabecalho(ws3, ['Número', 'Status', 'Cliente', 'Veículo', 'Placa', 'Mecânico', 'KM Entrada', 'Total Serviços', 'Total Peças', 'Desconto', 'Total Geral', 'Data Entrada'])
    for o in OrdemServico.objects.filter(oficina=oficina).select_related('cliente', 'veiculo', 'mecanico').prefetch_related('servicos', 'pecas_usadas').order_by('-data_entrada'):
        ws3.append([
            o.numero, o.get_status_display(), o.cliente.nome,
            f'{o.veiculo.marca} {o.veiculo.modelo}', o.veiculo.placa,
            o.mecanico.nome if o.mecanico else '', o.quilometragem_entrada,
            float(o.total_servicos), float(o.total_pecas), float(o.desconto), float(o.total_geral),
            o.data_entrada.strftime('%d/%m/%Y %H:%M') if o.data_entrada else '',
        ])

    ws4 = wb.create_sheet('Estoque')
    cabecalho(ws4, ['Código', 'Nome', 'Marca', 'Unidade', 'Quantidade', 'Qtd Mínima', 'Preço Custo', 'Preço Venda'])
    for p in Peca.objects.filter(oficina=oficina).order_by('nome'):
        ws4.append([p.codigo, p.nome, p.marca, p.unidade, float(p.quantidade), float(p.quantidade_minima), float(p.preco_custo), float(p.preco_venda)])

    ws5 = wb.create_sheet('Funcionários')
    cabecalho(ws5, ['ID', 'Nome', 'Telefone', 'Cargo', 'Ativo'])
    for f in Funcionario.objects.filter(oficina=oficina).order_by('nome'):
        ws5.append([f.id, f.nome, f.telefone, f.cargo, 'Sim' if f.ativo else 'Não'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="DoMecanico_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return resp
